import openpyxl as pyxl
from datetime import datetime
import os




def outputExcel(data, pth):
    if os.path.exists(pth):
    # open the file
        wb = pyxl.load_workbook(pth, data_only=True)
        sheet = wb.active #默認為最後一個sheet
    else:
    # no file, make new file
        now = datetime.now()
        now = now.strftime("%Y-%m-%d_%H-%M")
        wb = pyxl.Workbook()
        wb.save(pth)
        sheet = wb.active
        sheet.title = now
        
        sheet['A1'].value = '時間'
        sheet['B1'].value = '來源'
        sheet['C1'].value = '標題'
        sheet['D1'].value = '內文'
    
    # sheet = wb.activet
    idx = sheet.max_row+1 #last row in sheet

    for dic in data:
        sheet[f"A{idx}"] = dic['Time']
        sheet[f"B{idx}"] = dic['Source']
        sheet[f"C{idx}"] = dic['Title']
        sheet[f"D{idx}"] = dic['Content']
        idx+=1

        

    wb.save(pth)
    print(f"File saved：{pth}")